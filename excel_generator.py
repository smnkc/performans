import os
import random
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Kriterler ve numaraları (Varsayılan değerler, metin dosyaları yoksa oluşturulur)
PERFORMANCE_CRITERIA = [
    "Derse hazırlıklı gelme",
    "Öğretim materyallerini bulundurma",
    "Hazırbulunuşluluk düzeyi",
    "Öğrenme öğretme sürecine katılma",
    "Öğrenme öğretme sürecinde notlar alma",
    "Soru ve önerilere cevap verebilme",
    "Fikir yürütme, çıkarımda bulunma",
    "Tahmin ve gözlem yapabilme",
    "Grupla çalışma becerisi",
    "Türkçeyi güzel yazma ve doğru kullanma",
    "Derse karşı tutum (istekli oluş)",
    "Arkadaşlarına gösterdiği saygı",
    "Sınıf içi tartışmalara katılım",
    "Öğrenme öğretme sürecinde soru sorabilme",
    "Verilen görevleri yapabilme",
    "Konuları günlük yaşamla ilişkilendirme",
    "Eleştirel düşünme",
    "Analiz ve sentez yapabilme",
    "Etkinliklerde görev alma",
    "Yaratıcı düşünme becerisi"
]

PERFORMANCE_NUMBERS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]

PROJECT_CRITERIA = [
    "Projenin amacını belirleme",
    "Projeye uygun çalışma planı yapma",
    "Grup içinde görev dağılımı yapma (Grup çalışması ise)",
    "Belirlenen konunun önemini ortaya koyma",
    "Hazırlanan proje sonunda ne tür sonuçlara ulaşılmak istendiğini ortaya koyma",
    "Proje konusunda bilimsel açıdan doğru bilgiler aktarma",
    "Toplanan bilgileri analiz etme",
    "Elde edilen bilgilerden çıkarımda bulunma",
    "Yapılan çalışmanın orijinal olmasına özen gösterme",
    "Yapılan çıkarımların nedenlerini ortaya koyma",
    "Yapılan çalışmada eleştirel düşünme becerisini gösterme",
    "Hazırlanan raporun, resimler, gazete haberleri, çizimler, tablo, grafik ve istatistiklerle destekleme",
    "Metne aktarılan tüm bilgilerde Türkçeyi doğru biçimde kullanma",
    "Yararlanılan kaynakları rapora yansıtma",
    "Konuyu dinleyicilerin ilgisini çekecek şekilde sunma",
    "Sunuyu hedefe yönelik materyalle destekleme",
    "Sunuda akıcı bir dil ve beden dilini kullanma",
    "Sorulara cevap verme",
    "Verilen sürede sunuyu yapma",
    "Sunum sırasında Türkçeyi doğru biçimde kullanma"
]

def load_criteria_files():
    """
    Kriterleri klasördeki text dosyalarından yükler.
    Dosyalar yoksa varsayılan listelerle dosyaları otomatik oluşturur.
    """
    global PERFORMANCE_CRITERIA, PROJECT_CRITERIA
    perf_file = "performans_kriterleri.txt"
    proj_file = "proje_kriterleri.txt"
    
    # Performans Kriterleri Yükleme/Oluşturma
    if os.path.exists(perf_file):
        try:
            with open(perf_file, "r", encoding="utf-8") as f:
                lines = [line.strip() for line in f if line.strip()]
                if lines:
                    PERFORMANCE_CRITERIA.clear()
                    PERFORMANCE_CRITERIA.extend(lines)
        except Exception:
            pass
    else:
        try:
            with open(perf_file, "w", encoding="utf-8") as f:
                f.write("\n".join(PERFORMANCE_CRITERIA))
        except Exception:
            pass
            
    # Proje Kriterleri Yükleme/Oluşturma
    if os.path.exists(proj_file):
        try:
            with open(proj_file, "r", encoding="utf-8") as f:
                lines = [line.strip() for line in f if line.strip()]
                if lines:
                    PROJECT_CRITERIA.clear()
                    PROJECT_CRITERIA.extend(lines)
        except Exception:
            pass
    else:
        try:
            with open(proj_file, "w", encoding="utf-8") as f:
                f.write("\n".join(PROJECT_CRITERIA))
        except Exception:
            pass

def distribute_score(score, student_no, index_key, num_criteria=20):
    """
    Bir öğrencinin genel notunu (0-100) kriterlere dengeli şekilde dağıtır.
    Aynı öğrenci ve sınav için her zaman aynı dağılımı üretmesi için deterministik seed kullanır.
    """
    if score is None or score == "" or str(score).strip() == "":
        return [""] * num_criteria
        
    score_str = str(score).strip().upper()
    if score_str in ["G", "K"]:
        return [score_str] * num_criteria
        
    try:
        score_val = int(score_str)
    except ValueError:
        return [""] * num_criteria
        
    if score_val < 0:
        score_val = 0
    elif score_val > 100:
        score_val = 100
        
    # Dengeli dağıtım hesabı (ortalama üzerinden)
    base = score_val // num_criteria
    remainder = score_val % num_criteria
    
    scores = [base] * num_criteria
    for i in range(remainder):
        scores[i] += 1
        
    # Her öğrenciye özel deterministik karıştırma
    seed_str = f"{student_no}_{index_key}"
    try:
        seed_val = int(student_no) + hash(index_key)
    except ValueError:
        seed_val = hash(seed_str)
        
    rng = random.Random(seed_val)
    rng.shuffle(scores)
    
    return scores

def style_range(ws, cell_range, border=None, fill=None, font=None, alignment=None):
    """
    Belirli bir hücre aralığındaki tüm hücrelere stilleri uygular (birleştirilmiş hücreler için önemlidir).
    """
    for row in ws[cell_range]:
        for cell in row:
            if border:
                cell.border = border
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment

def create_excel_report(data, config, file_path, eval_type="Performans"):
    """
    Verileri çekilen e-Okul notlarına ve konfigürasyona göre Excel dosyası oluşturur.
    """
    # Klasördeki kriter dosyalarını yükle/yarat
    load_criteria_files()
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Stil tanımlamaları
    font_family = "Arial"
    font_title = Font(name=font_family, size=11, bold=True)
    font_header = Font(name=font_family, size=9, bold=True)
    font_bold = Font(name=font_family, size=9, bold=True)
    font_regular = Font(name=font_family, size=9, bold=False)
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_vertical = Alignment(textRotation=90, wrapText=True, horizontal='center', vertical='center')
    
    fill_grey = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    fill_header = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    
    border_thin_side = Side(style='thin', color='A0A0A0')
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    
    for class_key, class_data in data.items():
        sheet_name = class_key.replace(":", " ").replace("\\", " ").replace("/", " ").replace("?", " ").replace("*", " ")[:30]
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True
        
        # Sayfa Yapısı: Yatay (Landscape), A4 kağıt boyutu ve tek sayfaya sığdırma
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9  # A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        
        school_name = class_data.get("school_name", "").upper()
        academic_year = class_data.get("academic_year", "")
        term = class_data.get("term", "")
        program = class_data.get("program", "")
        course_name = class_data.get("course_name", "").upper()
        
        title_text = f"{school_name} {academic_year} {term} {program} {course_name} "
        if eval_type == "Performans":
            title_text += "DERS VE ETKİNLİKLERE KATILIM CETVELİ"
        else:
            title_text += "PROJE DEĞERLENDİRME ÖLÇEĞİ CETVELİ"
            
        students = class_data.get("students", [])
        
        if eval_type == "Performans":
            num_crit = len(PERFORMANCE_CRITERIA)
            # Toplam sütun sayısı: A,B,C (3) + Kriter Sayısı * 3 + 3 (toplamlar)
            max_col = 6 + num_crit * 3
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.cell(row=1, column=1, value=title_text)
            style_range(ws, f"A1:{get_column_letter(max_col)}1", font=font_title, alignment=align_center)
            ws.row_dimensions[1].height = 35
            
            # Öğrencinin bölümü
            ws.merge_cells("B2:C2")
            ws.cell(row=2, column=2, value="ÖĞRENCİNİN")
            style_range(ws, "B2:C2", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            # SIRA, NO, ADI SOYADI dikey birleştirme
            ws.merge_cells("A3:A6")
            ws.cell(row=3, column=1, value="SIRA")
            style_range(ws, "A3:A6", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            ws.merge_cells("B3:B6")
            ws.cell(row=3, column=2, value="NO")
            style_range(ws, "B3:B6", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            ws.merge_cells("C3:C6")
            ws.cell(row=3, column=3, value="ADI ve SOYADI")
            style_range(ws, "C3:C6", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            # Kriterleri ekle
            for i, crit in enumerate(PERFORMANCE_CRITERIA):
                col_start = 4 + i * 3
                col_letter_start = get_column_letter(col_start)
                col_letter_end = get_column_letter(col_start+2)
                
                # Kriter adı birleştirme (Row 2-3)
                ws.merge_cells(start_row=2, start_column=col_start, end_row=3, end_column=col_start+2)
                ws.cell(row=2, column=col_start, value=crit)
                style_range(ws, f"{col_letter_start}2:{col_letter_end}3", border=border_thin, fill=fill_header, font=Font(name=font_family, size=8, bold=True), alignment=align_vertical)
                
                # Numarası (Row 5) - Eğer kriter sayısı varsayılansa PERFORMANCE_NUMBERS kullan, değilse i+1 veya 10 yaz
                number_val = PERFORMANCE_NUMBERS[i] if i < len(PERFORMANCE_NUMBERS) else 10
                ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start+2)
                ws.cell(row=5, column=col_start, value=number_val)
                style_range(ws, f"{col_letter_start}5:{col_letter_end}5", border=border_thin, font=font_header, alignment=align_center)
                
                # I. II. III. sütunları (Row 6)
                ws.cell(row=6, column=col_start, value="I.").font = font_header
                ws.cell(row=6, column=col_start).alignment = align_center
                ws.cell(row=6, column=col_start).border = border_thin
                
                ws.cell(row=6, column=col_start+1, value="II.").font = font_header
                ws.cell(row=6, column=col_start+1).alignment = align_center
                ws.cell(row=6, column=col_start+1).border = border_thin
                
                c3_header = ws.cell(row=6, column=col_start+2, value="III.")
                c3_header.font = font_header
                c3_header.alignment = align_center
                c3_header.fill = fill_grey
                c3_header.border = border_thin
                
            # PUANLAR Merged (Row 4, Col D to last criteria col)
            last_crit_col = 3 + num_crit * 3
            last_crit_col_letter = get_column_letter(last_crit_col)
            ws.merge_cells(f"D4:{last_crit_col_letter}4")
            ws.cell(row=4, column=4, value="PUANLAR")
            style_range(ws, f"D4:{last_crit_col_letter}4", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            # Toplam sütunları başlıkları (BL, BM, BN -> dinamikleşti)
            total_1_idx = last_crit_col + 1
            total_2_idx = last_crit_col + 2
            total_3_idx = last_crit_col + 3
            
            t1_let = get_column_letter(total_1_idx)
            t2_let = get_column_letter(total_2_idx)
            t3_let = get_column_letter(total_3_idx)
            
            ws.merge_cells(f"{t1_let}2:{t1_let}6")
            ws.cell(row=2, column=total_1_idx, value="I. PERF.\nTOPLAM")
            style_range(ws, f"{t1_let}2:{t1_let}6", border=border_thin, fill=fill_header, font=font_header, alignment=align_vertical)
            
            ws.merge_cells(f"{t2_let}2:{t2_let}6")
            ws.cell(row=2, column=total_2_idx, value="II. PERF.\nTOPLAM")
            style_range(ws, f"{t2_let}2:{t2_let}6", border=border_thin, fill=fill_header, font=font_header, alignment=align_vertical)
            
            ws.merge_cells(f"{t3_let}2:{t3_let}6")
            ws.cell(row=2, column=total_3_idx, value="III. PERF.\nTOPLAM")
            style_range(ws, f"{t3_let}2:{t3_let}6", border=border_thin, fill=fill_header, font=font_header, alignment=align_vertical)
            
            ws.cell(row=2, column=1).border = border_thin
            ws.cell(row=1, column=1).border = border_thin
            
            ws.row_dimensions[2].height = 55
            ws.row_dimensions[3].height = 55
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 20
            
            # Öğrencileri yerleştir
            current_row = 7
            for idx, student in enumerate(students):
                ws.cell(row=current_row, column=1, value=idx+1).font = font_regular
                ws.cell(row=current_row, column=1).alignment = align_center
                
                ws.cell(row=current_row, column=2, value=student.get("no", "")).font = font_regular
                ws.cell(row=current_row, column=2).alignment = align_center
                
                ws.cell(row=current_row, column=3, value=student.get("name", "")).font = font_regular
                ws.cell(row=current_row, column=3).alignment = align_left
                
                # Performans puanlarını dinamik kriter sayısına göre dağıt
                perf1_list = distribute_score(student.get("perf1"), student.get("no"), "perf1", num_criteria=num_crit)
                perf2_list = distribute_score(student.get("perf2"), student.get("no"), "perf2", num_criteria=num_crit)
                perf3_list = distribute_score(student.get("perf3"), student.get("no"), "perf3", num_criteria=num_crit)
                
                for i in range(num_crit):
                    col_start = 4 + i * 3
                    # I. Perf
                    ws.cell(row=current_row, column=col_start, value=perf1_list[i]).font = font_regular
                    ws.cell(row=current_row, column=col_start).alignment = align_center
                    
                    # II. Perf
                    ws.cell(row=current_row, column=col_start+1, value=perf2_list[i]).font = font_regular
                    ws.cell(row=current_row, column=col_start+1).alignment = align_center
                    
                    # III. Perf (Gri Sütun)
                    c3 = ws.cell(row=current_row, column=col_start+2, value=perf3_list[i])
                    c3.font = font_regular
                    c3.alignment = align_center
                    c3.fill = fill_grey
                    
                # Formüllerle toplam alanları
                sum_cols_1 = ",".join([f"{get_column_letter(4 + i*3)}{current_row}" for i in range(num_crit)])
                ws.cell(row=current_row, column=total_1_idx, value=f"=SUM({sum_cols_1})").font = font_bold
                ws.cell(row=current_row, column=total_1_idx).alignment = align_center
                
                sum_cols_2 = ",".join([f"{get_column_letter(5 + i*3)}{current_row}" for i in range(num_crit)])
                ws.cell(row=current_row, column=total_2_idx, value=f"=SUM({sum_cols_2})").font = font_bold
                ws.cell(row=current_row, column=total_2_idx).alignment = align_center
                
                sum_cols_3 = ",".join([f"{get_column_letter(6 + i*3)}{current_row}" for i in range(num_crit)])
                ws.cell(row=current_row, column=total_3_idx, value=f"=SUM({sum_cols_3})").font = font_bold
                ws.cell(row=current_row, column=total_3_idx).alignment = align_center
                
                # Kenarlıklar
                for col in range(1, max_col+1):
                    ws.cell(row=current_row, column=col).border = border_thin
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
        else: # Proje Seçilmişse
            num_crit = len(PROJECT_CRITERIA)
            # Toplam sütun sayısı: A,B,C (3) + Kriter Sayısı + 1 (toplam)
            max_col = 4 + num_crit
            
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
            ws.cell(row=1, column=1, value=title_text)
            style_range(ws, f"A1:{get_column_letter(max_col)}1", font=font_title, alignment=align_center)
            ws.row_dimensions[1].height = 35
            
            # SIRA, NO, ADI SOYADI dikey birleştirme (Row 2-4)
            ws.merge_cells("A2:A4")
            ws.cell(row=2, column=1, value="SIRA")
            style_range(ws, "A2:A4", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            ws.merge_cells("B2:B4")
            ws.cell(row=2, column=2, value="NO")
            style_range(ws, "B2:B4", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            ws.merge_cells("C2:C4")
            ws.cell(row=2, column=3, value="ADI ve SOYADI")
            style_range(ws, "C2:C4", border=border_thin, fill=fill_header, font=font_header, alignment=align_center)
            
            # Proje kriter başlıkları (D2 to last criteria col)
            for i, crit in enumerate(PROJECT_CRITERIA):
                col_idx = 4 + i
                col_letter = get_column_letter(col_idx)
                ws.merge_cells(start_row=2, start_column=col_idx, end_row=4, end_column=col_idx)
                ws.cell(row=2, column=col_idx, value=crit)
                style_range(ws, f"{col_letter}2:{col_letter}4", border=border_thin, fill=fill_header, font=Font(name=font_family, size=8, bold=True), alignment=align_vertical)
                
            # Proje Toplam başlığı
            total_idx = 4 + num_crit
            t_let = get_column_letter(total_idx)
            ws.merge_cells(f"{t_let}2:{t_let}4")
            ws.cell(row=2, column=total_idx, value="PROJE\nTOPLAMI")
            style_range(ws, f"{t_let}2:{t_let}4", border=border_thin, fill=fill_header, font=font_header, alignment=align_vertical)
            
            ws.row_dimensions[2].height = 45
            ws.row_dimensions[3].height = 45
            ws.row_dimensions[4].height = 45
            
            # Öğrencileri yerleştir
            current_row = 5
            for idx, student in enumerate(students):
                ws.cell(row=current_row, column=1, value=idx+1).font = font_regular
                ws.cell(row=current_row, column=1).alignment = align_center
                
                ws.cell(row=current_row, column=2, value=student.get("no", "")).font = font_regular
                ws.cell(row=current_row, column=2).alignment = align_center
                
                ws.cell(row=current_row, column=3, value=student.get("name", "")).font = font_regular
                ws.cell(row=current_row, column=3).alignment = align_left
                
                proj_list = distribute_score(student.get("project"), student.get("no"), "project", num_criteria=num_crit)
                
                for i in range(num_crit):
                    col_idx = 4 + i
                    ws.cell(row=current_row, column=col_idx, value=proj_list[i]).font = font_regular
                    ws.cell(row=current_row, column=col_idx).alignment = align_center
                    
                # Proje Toplam Formülü
                ws.cell(row=current_row, column=total_idx, value=f"=SUM(D{current_row}:{get_column_letter(3 + num_crit)}{current_row})").font = font_bold
                ws.cell(row=current_row, column=total_idx).alignment = align_center
                
                # Kenarlıklar
                for col in range(1, max_col+1):
                    ws.cell(row=current_row, column=col).border = border_thin
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
        # Başlık ve sütun genişliklerini düzenleyelim
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        
        # Kriter genişliklerini ayarla (dar sütunlar)
        if eval_type == "Performans":
            for col in range(4, 4 + num_crit * 3):
                ws.column_dimensions[get_column_letter(col)].width = 4
            ws.column_dimensions[t1_let].width = 8
            ws.column_dimensions[t2_let].width = 8
            ws.column_dimensions[t3_let].width = 8
        else:
            for col in range(4, 4 + num_crit):
                ws.column_dimensions[get_column_letter(col)].width = 4.5
            ws.column_dimensions[t_let].width = 9
            
        # Alt İmza Blokları
        sig_row_start = current_row + 2
        
        # Sol taraf öğretmen bilgisi
        teacher_name = config.get("teacher_name", "")
        branch = config.get("branch", "")
        
        # Sağ taraf müdür bilgisi
        principal_name = config.get("principal_name", "")
        
        # Yazıları yerleştirelim
        # Sol Taraf
        ws.cell(row=sig_row_start, column=3, value=teacher_name).font = font_bold
        ws.cell(row=sig_row_start, column=3).alignment = align_left
        ws.cell(row=sig_row_start+1, column=3, value=branch).font = font_bold
        ws.cell(row=sig_row_start+1, column=3).alignment = align_left
        
        # Sağ Taraf (Tablonun en sağına hizalanacak şekilde)
        principal_col = max_col - 8 if max_col > 12 else 3
        if principal_col < 4:
            principal_col = max_col - 2
            
        ws.merge_cells(start_row=sig_row_start, start_column=principal_col, end_row=sig_row_start, end_column=max_col)
        ws.cell(row=sig_row_start, column=principal_col, value=principal_name).font = font_bold
        ws.cell(row=sig_row_start, column=principal_col).alignment = align_center
        
        ws.merge_cells(start_row=sig_row_start+1, start_column=principal_col, end_row=sig_row_start+1, end_column=max_col)
        ws.cell(row=sig_row_start+1, column=principal_col, value="OKUL MÜDÜRÜ").font = font_bold
        ws.cell(row=sig_row_start+1, column=principal_col).alignment = align_center
        
    wb.save(file_path)
    return True
